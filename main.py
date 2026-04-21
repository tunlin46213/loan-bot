from dotenv import load_dotenv
import os
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from keep_alive import keep_alive
from openai import OpenAI
from upstash_redis import Redis
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update, ReplyKeyboardRemove, BotCommand, BotCommandScopeChat
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    filters,
    ConversationHandler,
    CallbackQueryHandler,
    ContextTypes
)

(
    METHOD, AMOUNT, RATE, TERM_MONTHS, 
    VAL_DISTRICT, VAL_SIZE,
    SCORE_INCOME, SCORE_DEBT, SCORE_PROP_VALUE, SCORE_LOAN_AMOUNT,
    ADMIN_MENU, ADMIN_REVOKE_INPUT, ADMIN_ADD_INPUT, ADMIN_BROADCAST_INPUT
) = range(14)

# Load .env
load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN")
DASHSCOPE_API_KEY = os.getenv("DASHSCOPE_API_KEY")
ACCESS_CODE = os.getenv("ACCESS_CODE", "neat17112024")
ADMIN_ID = int(os.getenv("ADMIN_ID", "0"))

# Upstash Redis Client (Persistent Storage)
redis_client = Redis(
    url=os.getenv("UPSTASH_REDIS_REST_URL"),
    token=os.getenv("UPSTASH_REDIS_REST_TOKEN")
)

# Qwen Client
qwen_client = OpenAI(
    api_key=DASHSCOPE_API_KEY,
    base_url="https://openrouter.ai/api/v1"
)

SYSTEM_PROMPT = """
You are a BRED Bank Cambodia real estate loan expert assistant for bank staff.
Always base your answers and calculations exactly on BRED Bank's current policies:
- Maximum LTV: up to 80% (Requires 20% down payment)
- Maximum Loan Term: 20 years (240 months)
- Interest Rates: Years 1-5 at 8.50% p.a., Years 6-15 at 8.75% p.a., Years 16-20 at 8.95% p.a.
- DSR Requirement: Generally between 40% and 50%
- Fees: No loan approval fee
- Collateral: Hard title property is required
- Non-offered Loans: BRED Bank Cambodia DOES NOT offer car loans or any vehicle financing. Only real estate (property) loans are offered.

When explaining calculations like Loan-to-Value (LTV) or Debt Service Ratio (DSR), DO NOT use complex math equations or LaTeX.
Instead, use plain text and very simple step-by-step explanations.
For example: 'Loan Amount ÷ Property Value = LTV'
Use simple analogies and make it easy to understand for beginners.
Always reply in the exact same language (e.g. Burmese, English, etc.) as the user's message.
"""

user_conversations = {}

def get_main_menu(user_id=None):
    # Use Telegram native Menu button — no custom reply keyboard
    return ReplyKeyboardRemove()

async def start(update, context):
    try:
        user_id = update.effective_user.id
        is_authed = redis_client.sismember("auth_users", str(user_id))
    except Exception:
        is_authed = False
    
    if is_authed:
        await update.message.reply_text(
            "🏦 **BRED Bank Cambodia — Loan Tool**\n\n"
            "Welcome back! Use the Menu button below to get started.",
            reply_markup=ReplyKeyboardRemove(),
            parse_mode='Markdown'
        )
    else:
        await update.message.reply_text(
            "🏦 **BRED Bank Cambodia — Loan Tool**\n\n"
            "This system is for BRED Bank staff only.\n"
            "Please enter the access password to continue.\n\n"
            "Available commands:\n"
            "`/calculator` — Loan Amortization Calculator\n"
            "`/score`      — Loan Pre-approval Scoring\n"
            "`/valuation`  — Property Valuation Tool\n"
            "`/myid`       — View your Telegram ID",
            parse_mode='Markdown'
        )
    return ConversationHandler.END

async def handle_message(update, context):
    if not update.message or not update.message.text:
        return

    user_id = update.effective_user.id
    user_text = update.message.text

    # --- Access Control Check ---
    try:
        is_authed = redis_client.sismember("auth_users", str(user_id))
    except Exception as e:
        print(f"Redis error in handle_message: {e}")
        is_authed = False

    if not is_authed:
        if user_text.strip() == ACCESS_CODE:
            try:
                redis_client.sadd("auth_users", str(user_id))
                # Save user info for admin panel
                user = update.effective_user
                redis_client.hset(f"user_info:{user_id}", mapping={
                    "name": user.full_name or "Unknown",
                    "username": user.username or ""
                })
            except Exception as e:
                print(f"Redis error saving user: {e}")
            await update.message.reply_text(
                "✅ Access granted! Use the Menu button to get started.",
                reply_markup=ReplyKeyboardRemove()
            )
        else:
            await update.message.reply_text("🔒 This bot is restricted. Please enter the correct password:")
        return
    # ----------------------------

    if user_id not in user_conversations:
        user_conversations[user_id] = []

    user_conversations[user_id].append({
        "role": "user",
        "content": user_text
    })

    await context.bot.send_chat_action(
        chat_id=update.effective_chat.id,
        action="typing"
    )

    response = qwen_client.chat.completions.create(
        model="qwen/qwen-plus",
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            *user_conversations[user_id]
        ]
    )

    bot_reply = response.choices[0].message.content

    user_conversations[user_id].append({
        "role": "assistant",
        "content": bot_reply
    })

    # Keep last 10 messages only
    if len(user_conversations[user_id]) > 10:
        user_conversations[user_id] = \
            user_conversations[user_id][-10:]

    await update.message.reply_text(bot_reply)

# --- Enterprise Calculator Functions ---
async def start_calculator(update, context):
    try:
        is_authed = redis_client.sismember("auth_users", str(update.effective_user.id))
    except Exception:
        is_authed = False
    if not is_authed:
        if update.message:
            await update.message.reply_text("🔒 Please enter the password before using the calculator.")
        return ConversationHandler.END
        
    keyboard = [
        [InlineKeyboardButton("Principal + Interest Equal (EMI)", callback_data='emi')],
        [InlineKeyboardButton("Principal Equal Payment", callback_data='equal_principal')],
        [InlineKeyboardButton("Principal Bullet Repayment", callback_data='bullet')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    msg = "🔢 **Enterprise Loan Calculator**\n\nPlease select the calculation method:"
    if update.message:
        await update.message.reply_text(msg, reply_markup=reply_markup, parse_mode='Markdown')
    return METHOD

async def select_method(update, context):
    query = update.callback_query
    await query.answer()
    context.user_data['method'] = query.data
    
    method_name = ""
    if query.data == 'emi':
        method_name = "Principal + Interest Equal Payment"
    elif query.data == 'equal_principal':
        method_name = "Principal Equal Payment"
    elif query.data == 'bullet':
        method_name = "Principal Bullet Repayment"
        
    await query.edit_message_text(f"✅ Selected: **{method_name}**\n\n💰 Please enter the **Loan Amount** in USD (e.g. 100000):", parse_mode='Markdown')
    return AMOUNT

async def get_amount(update, context):
    try:
        text = update.message.text.replace(',', '').replace('$', '')
        context.user_data['amount'] = float(text)
        await update.message.reply_text("📈 Enter the **Annual Interest Rate** in % (e.g. 8.5):", parse_mode='Markdown')
        return RATE
    except ValueError:
        await update.message.reply_text("❌ Invalid format. Please enter a number for Loan Amount (e.g. 100000):")
        return AMOUNT

async def get_rate(update, context):
    try:
        text = update.message.text.replace('%', '')
        context.user_data['rate'] = float(text)
        await update.message.reply_text("⏳ Enter the **Loan Term in Months** (e.g. 54, 368):", parse_mode='Markdown')
        return TERM_MONTHS
    except ValueError:
        await update.message.reply_text("❌ Invalid format. Please enter a number for Interest Rate (e.g. 8.5):")
        return RATE

async def get_term_months(update, context):
    try:
        months = int(update.message.text)
        amount = context.user_data['amount']
        rate = context.user_data['rate']
        method = context.user_data['method']

        # --- Method label ---
        method_labels = {
            'emi':             'Principal + Interest Equal (EMI)',
            'equal_principal': 'Principal Equal Payment',
            'bullet':          'Principal Bullet Repayment',
        }
        method_label = method_labels.get(method, method)

        balance = amount
        monthly_rate = (rate / 100) / 12
        start_date = date.today()

        # --- Pre-calculate scheduled payment for header ---
        if method == 'emi':
            if monthly_rate > 0:
                scheduled_payment = amount * monthly_rate * ((1 + monthly_rate)**months) / (((1 + monthly_rate)**months) - 1)
            else:
                scheduled_payment = amount / months
            actual_payments = months
        elif method == 'equal_principal':
            principal_payment = amount / months
            scheduled_payment = principal_payment + (amount * monthly_rate)  # First payment (highest)
            actual_payments = months
        elif method == 'bullet':
            scheduled_payment = amount * monthly_rate  # Monthly interest only
            actual_payments = months
        else:
            scheduled_payment = 0
            actual_payments = months

        # --- Build amortization schedule rows ---
        schedule_rows = []
        total_payment = 0
        total_interest = 0
        total_principal = 0
        cumulative_interest = 0
        balance = amount

        for i in range(1, months + 1):
            beginning_balance = balance
            payment_date = (start_date + relativedelta(months=i)).strftime('%m/%d/%Y')
            interest = beginning_balance * monthly_rate
            extra_payment = 0.0

            if method == 'emi':
                sched_pay = scheduled_payment
                principal = sched_pay - interest
                if i == months:  # Final payment adjustment
                    principal = beginning_balance
                    sched_pay = principal + interest
            elif method == 'equal_principal':
                principal = amount / months
                sched_pay = principal + interest
            elif method == 'bullet':
                if i == months:
                    principal = beginning_balance
                else:
                    principal = 0.0
                sched_pay = principal + interest
            else:
                principal = 0
                sched_pay = interest

            total_pay = round(sched_pay + extra_payment, 2)
            principal = round(principal, 2)
            interest = round(interest, 2)
            sched_pay = round(sched_pay, 2)
            ending_balance = round(max(0.0, beginning_balance - principal), 2)
            cumulative_interest = round(cumulative_interest + interest, 2)

            schedule_rows.append([
                i,
                payment_date,
                round(beginning_balance, 2),
                round(sched_pay, 2),
                round(extra_payment, 2),
                round(total_pay, 2),
                round(principal, 2),
                round(interest, 2),
                round(ending_balance, 2),
                round(cumulative_interest, 2),
            ])

            balance = ending_balance
            total_payment += total_pay
            total_interest += interest
            total_principal += principal

        first_payment_num = schedule_rows[0][3] if schedule_rows else 0.0
        first_payment = f"${first_payment_num:,.2f}"

        # --- Khmer method labels ---
        method_labels_khmer = {
            'emi':             'ដើម + ការប្រាក់ស្មើគ្នា (EMI)',
            'equal_principal': 'ដើមស្មើគ្នា',
            'bullet':          'ការប្រាក់ប្រចាំខែ / ដើមចុងក្រោយ',
        }
        method_label_kh = method_labels_khmer.get(method, method_label)

        # ── Styles ────────────────────────────────────────────
        C_DARK_BLUE  = "1F3864"
        C_MED_BLUE   = "2E75B6"
        C_LIGHT_BLUE = "D9E2F3"
        C_ALT_BLUE   = "BDD7EE"
        C_ORANGE_TXT = "C45911"
        C_ORANGE_BG  = "F4B183"
        C_WHITE      = "FFFFFF"

        f_title  = Font(name="Calibri", bold=True,  size=11, underline="single")
        f_hdr    = Font(name="Calibri", bold=True,  size=10, color=C_WHITE)
        f_label  = Font(name="Calibri",             size=10, color="404040")
        f_val    = Font(name="Calibri", bold=True,  size=10, color=C_ORANGE_TXT)
        f_data   = Font(name="Calibri",             size=10)
        f_total  = Font(name="Calibri", bold=True,  size=10)
        f_lender = Font(name="Calibri", bold=True,  size=11)
        f_method = Font(name="Calibri", bold=True, italic=True, size=10, color=C_MED_BLUE)

        fl_hdr    = PatternFill("solid", fgColor=C_DARK_BLUE)
        fl_orange = PatternFill("solid", fgColor=C_ORANGE_BG)
        fl_even   = PatternFill("solid", fgColor=C_LIGHT_BLUE)
        fl_total  = PatternFill("solid", fgColor=C_ALT_BLUE)

        def _thin(color="B8CCE4"):
            s = Side(style="thin", color=color)
            return Border(left=s, right=s, top=s, bottom=s)

        def _hdr_border():
            s = Side(style="medium", color=C_WHITE)
            return Border(left=s, right=s, top=s, bottom=s)

        a_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
        a_r = Alignment(horizontal="right",  vertical="center")
        a_l = Alignment(horizontal="left",   vertical="center")

        # ── Workbook ──────────────────────────────────────────
        wb = Workbook()
        ws = wb.active
        ws.title = "Loan Amortization"
        ws.sheet_properties.tabColor = "1F3864"   # BRED Bank blue tab

        col_widths = [6, 14, 18, 18, 14, 16, 14, 12, 16, 19]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        def put(r, c, v="", font=f_label, fill=None, align=a_l, border=None, num_fmt=None):
            cl = ws.cell(row=r, column=c, value=v)
            cl.font = font
            if fill:    cl.fill   = fill
            cl.alignment = align
            if border:  cl.border = border
            if num_fmt: cl.number_format = num_fmt
            return cl

        def merge(r, c1, c2):
            ws.merge_cells(start_row=r, start_column=c1,
                           end_row=r,   end_column=c2)

        # ══ SECTION 1 – Header ═══════════════════════════════
        row = 1
        ws.row_dimensions[row].height = 18
        merge(row, 1, 5); put(row, 1, "ENTER VALUES",  font=f_title)
        merge(row, 6,10); put(row, 6, "LOAN SUMMARY",  font=f_title)
        row += 1

        # Enter Values rows (left) + Loan Summary rows (right)
        e_rows = [
            ("Loan amount",               f"${amount:,.2f}"),
            ("Annual interest rate",      f"{rate:.2f}%"),
            ("Loan period in months",     str(months)),
            ("Number of payments / year", "12"),
            ("Start date of loan",        start_date.strftime("%m/%d/%Y")),
        ]
        s_rows = [
            ("Scheduled payment",            first_payment),
            ("Scheduled number of payments", str(months)),
            ("Actual number of payments",    str(actual_payments)),
            ("Total early payments",         "$0.00"),
            ("Total interest",               f"${total_interest:,.2f}"),
        ]
        for (el, ev), (sl, sv) in zip(e_rows, s_rows):
            ws.row_dimensions[row].height = 16
            merge(row, 1, 2); put(row, 1, el, border=_thin())
            merge(row, 3, 5); put(row, 3, ev, font=f_val, fill=fl_orange, align=a_r, border=_thin())
            merge(row, 6, 7); put(row, 6, sl, border=_thin())
            merge(row, 8,10); put(row, 8, sv, font=f_val, fill=fl_orange, align=a_r, border=_thin())
            row += 1

        # Optional extra payments
        ws.row_dimensions[row].height = 16
        merge(row, 1, 2); put(row, 1, "Optional extra payments", border=_thin())
        merge(row, 3, 5); put(row, 3, "$0.00", font=f_val, fill=fl_orange, align=a_r, border=_thin())
        row += 2  # blank gap

        # Lender Name + Repayment Method
        ws.row_dimensions[row].height = 18
        merge(row, 1, 2); put(row, 1, "LENDER NAME", font=f_lender, border=_thin())
        merge(row, 3, 5); put(row, 3, "BRED BANK Cambodia",
                               font=Font(name="Calibri", bold=True, size=11, color=C_MED_BLUE),
                               align=a_c, border=_thin())
        merge(row, 6, 7); put(row, 6, "Repayment Method", font=f_lender, border=_thin())
        merge(row, 8,10); put(row, 8, method_label, font=f_method, align=a_r, border=_thin())
        row += 2  # blank gap

        # ══ SECTION 2 – Table Header ══════════════════════════
        ws.row_dimensions[row].height = 36
        col_headers = [
            "PMT\nNO", "PAYMENT\nDATE", "BEGINNING\nBALANCE",
            "SCHEDULED\nPAYMENT", "EXTRA\nPAYMENT", "TOTAL\nPAYMENT",
            "PRINCIPAL", "INTEREST", "ENDING\nBALANCE", "CUMULATIVE\nINTEREST"
        ]
        for c, h in enumerate(col_headers, 1):
            put(row, c, h, font=f_hdr, fill=fl_hdr, align=a_c, border=_hdr_border())
        table_hdr_row = row   # save for print_title_rows
        row += 1

        # Freeze panes below header + lender info
        ws.freeze_panes = ws.cell(row=row, column=1)

        # ══ SECTION 3 – Data Rows ═════════════════════════════
        MONEY_FMT = '$#,##0.00'
        for i, sr in enumerate(schedule_rows):
            fill = fl_even if i % 2 == 0 else None
            ws.row_dimensions[row].height = 15
            for c, val in enumerate(sr, 1):
                al  = a_c if c <= 2 else a_r
                nf  = MONEY_FMT if c >= 3 else None
                put(row, c, val, font=f_data, fill=fill, align=al, border=_thin(), num_fmt=nf)
            row += 1

        # ══ SECTION 4 – Totals Row ════════════════════════════
        ws.row_dimensions[row].height = 18
        merge(row, 1, 5); put(row, 1, "TOTAL", font=f_total, fill=fl_total, align=a_c, border=_thin())
        totals_vals = [
            (6,  total_payment,   MONEY_FMT),
            (7,  total_principal, MONEY_FMT),
            (8,  total_interest,  MONEY_FMT),
            (9,  "",              None),
            (10, "",              None),
        ]
        for c, v, nf in totals_vals:
            put(row, c, v, font=f_total, fill=fl_total, align=a_r, border=_thin(), num_fmt=nf)

        # ── Print Settings ────────────────────────────────────
        from openpyxl.worksheet.page import PageMargins
        ws.page_setup.orientation        = 'landscape'
        ws.page_setup.fitToPage          = True
        ws.page_setup.fitToWidth         = 1
        ws.page_setup.fitToHeight        = 0
        ws.page_setup.paperSize          = ws.PAPERSIZE_A4
        ws.page_margins                  = PageMargins(left=0.5, right=0.5, top=0.6, bottom=0.6)
        ws.print_title_rows              = f'1:{table_hdr_row}'   # repeat header on each page
        ws.print_area                    = f'A1:J{row}'
        ws.sheet_view.showGridLines      = True

        # ── Save ──────────────────────────────────────────────
        file_name = f"Loan_Schedule_{update.effective_user.id}.xlsx"
        wb.save(file_name)

        msg = (
            f"📊 **លទ្ធផលតារាងបង់ប្រាក់កម្ចី**\n"
            f"វិធីសាស្ត្រ: {method_label_kh}\n\n"
            f"💰 ចំនួនប្រាក់កម្ចី: `${amount:,.2f}`\n"
            f"📈 អត្រាការប្រាក់: `{rate}%`\n"
            f"⏳ រយៈពេល: `{months}` ខែ\n"
            f"📅 កាលបរិច្ឆេទចាប់ផ្ដើម: `{start_date.strftime('%m/%d/%Y')}`\n"
            f"─────────────────────────\n"
            f"💵 ការបង់ប្រាក់ប្រចាំខែ: `{first_payment}`\n"
            f"💵 ការប្រាក់សរុប: `${total_interest:,.2f}`\n"
            f"💵 ការទូទាត់សរុប: `${total_payment:,.2f}`\n\n"
            f"📄 តារាងបង់ប្រាក់ប្រចាំខែលម្អិតត្រូវបានភ្ជាប់ខាងលើ។"
        )

        with open(file_name, 'rb') as doc:
            await update.message.reply_document(
                document=doc,
                caption=msg,
                parse_mode='Markdown'
            )

        os.remove(file_name)
        return ConversationHandler.END

    except ValueError:
        await update.message.reply_text("❌ Invalid format. Please enter a whole number for Term in Months (e.g. 54):")
        return TERM_MONTHS

async def cancel_calculator(update, context):
    if update.message:
        await update.message.reply_text("❌ Calculator cancelled. You can ask me any loan questions.")
    return ConversationHandler.END

# --- Property Valuation Tool Functions ---
MEDIAN_PRICES = {
    'daun_penh': 5500,
    '7_makara': 6500,
    'bkk': 5000,
    'toul_kork': 4500,
    'chamkarmon': 3750,
    'sen_sok': 1400,
    'chroy_changvar': 1350,
    'kamboul': 300
}

async def start_valuation(update, context):
    try:
        is_authed = redis_client.sismember("auth_users", str(update.effective_user.id))
    except Exception:
        is_authed = False
    if not is_authed:
        if update.message: await update.message.reply_text("🔒 Please enter password.")
        return ConversationHandler.END
        
    keyboard = [
        [InlineKeyboardButton("Daun Penh", callback_data='daun_penh'), InlineKeyboardButton("7 Makara", callback_data='7_makara')],
        [InlineKeyboardButton("BKK", callback_data='bkk'), InlineKeyboardButton("Toul Kork", callback_data='toul_kork')],
        [InlineKeyboardButton("Chamkarmon", callback_data='chamkarmon'), InlineKeyboardButton("Sen Sok", callback_data='sen_sok')],
        [InlineKeyboardButton("Chroy Changvar", callback_data='chroy_changvar'), InlineKeyboardButton("Kamboul", callback_data='kamboul')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    if update.message:
        await update.message.reply_text("🏢 **Property Valuation Tool**\n\nSelect the District in Phnom Penh:", reply_markup=reply_markup, parse_mode='Markdown')
    return VAL_DISTRICT

async def select_district(update, context):
    query = update.callback_query
    await query.answer()
    context.user_data['val_district'] = query.data
    district_name = query.data.replace('_', ' ').title()
    await query.edit_message_text(f"✅ Selected: **{district_name}**\n\n📏 Please enter the **Property Size in Sqm** (e.g. 150):", parse_mode='Markdown')
    return VAL_SIZE

async def get_val_size(update, context):
    try:
        sqm = float(update.message.text.replace('sqm', '').strip())
        district = context.user_data['val_district']
        price_per_sqm = MEDIAN_PRICES[district]
        total_value = sqm * price_per_sqm
        district_name = district.replace('_', ' ').title()
        
        msg = (
            f"🏠 **Estimated Property Value**\n\n"
            f"📍 **Location:** {district_name}\n"
            f"📏 **Size:** {sqm} Sqm\n"
            f"💲 **Indicative Price:** ${price_per_sqm:,.2f} / Sqm\n"
            f"---------------------------------\n"
            f"💰 **Total Estimated Value:** ${total_value:,.2f}\n\n"
            f"*(Note: This is an indicative estimate based on 2024 median prices. Actual value varies by Sangkat and road access.)*"
        )
        await update.message.reply_text(msg, parse_mode='Markdown')
        return ConversationHandler.END
    except ValueError:
        await update.message.reply_text("❌ Invalid format. Please enter a valid number for Sqm (e.g. 150):")
        return VAL_SIZE

async def cancel_valuation(update, context):
    if update.message: await update.message.reply_text("❌ Valuation cancelled.")
    return ConversationHandler.END

# --- Loan Pre-approval Scoring Functions ---
async def start_score(update, context):
    try:
        is_authed = redis_client.sismember("auth_users", str(update.effective_user.id))
    except Exception:
        is_authed = False
    if not is_authed:
        if update.message: await update.message.reply_text("🔒 Please enter password.")
        return ConversationHandler.END
    if update.message:
        await update.message.reply_text("📋 **Loan Pre-approval Scoring**\n\n💵 Please enter the applicant's **Total Monthly Income** in USD (e.g. 3000):", parse_mode='Markdown')
    return SCORE_INCOME

async def get_score_income(update, context):
    try:
        context.user_data['score_income'] = float(update.message.text.replace(',', '').replace('$', ''))
        await update.message.reply_text("💳 Enter existing **Total Monthly Debts/Outgoings** in USD (e.g. 500):", parse_mode='Markdown')
        return SCORE_DEBT
    except ValueError:
        await update.message.reply_text("❌ Invalid number. Try again (e.g. 3000):")
        return SCORE_INCOME

async def get_score_debt(update, context):
    try:
        context.user_data['score_debt'] = float(update.message.text.replace(',', '').replace('$', ''))
        await update.message.reply_text("🏢 Enter the **Target Property Value** in USD (e.g. 150000):", parse_mode='Markdown')
        return SCORE_PROP_VALUE
    except ValueError:
        await update.message.reply_text("❌ Invalid number. Try again (e.g. 500):")
        return SCORE_DEBT

async def get_score_prop_value(update, context):
    try:
        context.user_data['score_prop_value'] = float(update.message.text.replace(',', '').replace('$', ''))
        await update.message.reply_text("💰 Enter the **Requested Loan Amount** in USD (e.g. 100000):", parse_mode='Markdown')
        return SCORE_LOAN_AMOUNT
    except ValueError:
        await update.message.reply_text("❌ Invalid number. Try again (e.g. 150000):")
        return SCORE_PROP_VALUE

async def get_score_loan_amount(update, context):
    try:
        loan_amount = float(update.message.text.replace(',', '').replace('$', ''))
        income = context.user_data['score_income']
        debt = context.user_data['score_debt']
        prop_value = context.user_data['score_prop_value']
        
        # Calculate LTV
        ltv = (loan_amount / prop_value) * 100 if prop_value > 0 else 0
        
        # Estimate new EMI roughly (assuming 8.5% over 20 years for a quick test)
        monthly_rate = (8.5 / 100) / 12
        months = 240
        if monthly_rate > 0:
            estimated_emi = loan_amount * monthly_rate * ((1 + monthly_rate)**months) / (((1 + monthly_rate)**months) - 1)
        else:
            estimated_emi = loan_amount / months
            
        total_debt_proposed = debt + estimated_emi
        dsr = (total_debt_proposed / income) * 100 if income > 0 else 100
        
        # Risk Assessment
        risk_level = "✅ **LOW RISK** (Likely Approved)"
        warnings = []
        
        if ltv > 80:
            risk_level = "❌ **HIGH RISK** (Likely Rejected)"
            warnings.append("- LTV exceeds BRED's 80% maximum limit. Need a larger down payment or higher property value.")
        if dsr > 50:
            risk_level = "❌ **HIGH RISK** (Likely Rejected)"
            warnings.append("- DSR exceeds 50%. Income is insufficient for this loan amount + existing debts.")
        elif dsr > 40:
            if risk_level != "❌ **HIGH RISK** (Likely Rejected)":
                risk_level = "⚠️ **MEDIUM RISK** (Needs Review)"
            warnings.append("- DSR is between 40% and 50%. This is borderline acceptable.")
            
        if not warnings:
            warnings.append("- Meets both LTV (<80%) and DSR (<40%) guidelines cleanly.")

        warn_text = "\n".join(warnings)
        
        msg = (
            f"📊 **Loan Pre-approval Scorecard**\n\n"
            f"**1. LTV (Loan-to-Value) Check:**\n"
            f"   - Loan: ${loan_amount:,.2f} | Property: ${prop_value:,.2f}\n"
            f"   - **LTV Ratio:** {ltv:.1f}%\n\n"
            f"**2. DSR (Debt Service Ratio) Check:**\n"
            f"   - Monthly Income: ${income:,.2f}\n"
            f"   - Existing Debts: ${debt:,.2f}\n"
            f"   - Estimated New EMI (20Y @ 8.5%): ${estimated_emi:,.2f}\n"
            f"   - **DSR Ratio:** {dsr:.1f}%\n\n"
            f"**Result:** {risk_level}\n"
            f"**Notes:**\n{warn_text}"
        )
        await update.message.reply_text(msg, parse_mode='Markdown')
        return ConversationHandler.END
        
    except ValueError:
        await update.message.reply_text("❌ Invalid number. Try again (e.g. 100000):")
        return SCORE_LOAN_AMOUNT

async def cancel_score(update, context):
    if update.message: await update.message.reply_text("❌ Scoring cancelled.")
    return ConversationHandler.END
# --------------------------------

# --- /myid Command ---
async def myid(update, context):
    user = update.effective_user
    await update.message.reply_text(
        f"🆔 **Your Telegram User ID:**\n\n`{user.id}`\n\n"
        f"Share this ID with the Admin to get access, or set it as ADMIN\\_ID in Render environment variables.",
        parse_mode='Markdown'
    )

# --- Admin Panel ---
async def admin_panel(update, context):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("❌ Access denied. You are not the admin.")
        return ConversationHandler.END
    user_count = redis_client.scard("auth_users")
    keyboard = [
        [InlineKeyboardButton("👥 View All Users", callback_data='admin_view')],
        [InlineKeyboardButton("❌ Revoke User", callback_data='admin_revoke'),
         InlineKeyboardButton("➕ Add User", callback_data='admin_add')],
        [InlineKeyboardButton("📢 Broadcast", callback_data='admin_broadcast')]
    ]
    await update.message.reply_text(
        f"👑 **Admin Panel**\n\n"
        f"📊 Total Authenticated Users: `{user_count}`\n\n"
        f"Select an action:",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='Markdown'
    )
    return ADMIN_MENU

async def admin_callback(update, context):
    query = update.callback_query
    await query.answer()
    if query.data == 'admin_view':
        members = redis_client.smembers("auth_users")
        if not members:
            await query.edit_message_text("📭 No authenticated users found.")
            return ConversationHandler.END
        user_list = ""
        for i, uid in enumerate(sorted(members), 1):
            info = redis_client.hgetall(f"user_info:{uid}")
            name = info.get("name", "Unknown") if info else "Unknown"
            username = info.get("username", "") if info else ""
            username_str = f" (@{username})" if username else ""
            user_list += f"`{i}.` `{uid}` — {name}{username_str}\n"
        await query.edit_message_text(
            f"👥 **Authenticated Users ({len(members)})**\n\n{user_list}\nUse /admin to go back.",
            parse_mode='Markdown'
        )
        return ConversationHandler.END
    elif query.data == 'admin_revoke':
        members = redis_client.smembers("auth_users")
        if not members:
            await query.edit_message_text("📭 No users to revoke.")
            return ConversationHandler.END
        keyboard = []
        for uid in sorted(members):
            info = redis_client.hgetall(f"user_info:{uid}")
            name = info.get("name", "Unknown") if info else "Unknown"
            username = info.get("username", "") if info else ""
            label = f"❌ {name} (@{username})" if username else f"❌ {name} [{uid}]"
            keyboard.append([InlineKeyboardButton(label, callback_data=f"revoke_{uid}")])
        keyboard.append([InlineKeyboardButton("⬅️ Back", callback_data="admin_back")])
        await query.edit_message_text(
            "❌ **Select user to revoke access:**",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='Markdown'
        )
        return ADMIN_REVOKE_INPUT
    elif query.data == 'admin_back':
        user_count = redis_client.scard("auth_users")
        keyboard = [
            [InlineKeyboardButton("👥 View All Users", callback_data='admin_view')],
            [InlineKeyboardButton("❌ Revoke User", callback_data='admin_revoke'),
             InlineKeyboardButton("➕ Add User", callback_data='admin_add')],
            [InlineKeyboardButton("📢 Broadcast", callback_data='admin_broadcast')]
        ]
        await query.edit_message_text(
            f"👑 **Admin Panel**\n\n"
            f"📊 Total Authenticated Users: `{user_count}`\n\n"
            f"Select an action:",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='Markdown'
        )
        return ADMIN_MENU
    elif query.data == 'admin_add':
        await query.edit_message_text(
            "➕ **Add User Access**\n\nEnter the **User ID** to grant access:\n_(Send /cancel to abort)_",
            parse_mode='Markdown'
        )
        return ADMIN_ADD_INPUT
    elif query.data == 'admin_broadcast':
        await query.edit_message_text(
            "📢 **Broadcast Message**\n\nType message to send to ALL users:\n_(Send /cancel to abort)_",
            parse_mode='Markdown'
        )
        return ADMIN_BROADCAST_INPUT

async def admin_revoke_callback(update, context):
    query = update.callback_query
    await query.answer()
    if query.data == 'admin_back':
        return ADMIN_MENU
    target_id = query.data.replace("revoke_", "")
    try:
        removed = redis_client.srem("auth_users", target_id)
        if removed:
            await query.edit_message_text(f"✅ User `{target_id}` access **revoked** successfully.", parse_mode='Markdown')
        else:
            await query.edit_message_text(f"⚠️ User `{target_id}` not found.", parse_mode='Markdown')
    except Exception as e:
        await query.edit_message_text(f"❌ Error: {e}")
    return ConversationHandler.END

async def admin_add_user(update, context):
    target_id = update.message.text.strip()
    try:
        redis_client.sadd("auth_users", target_id)
        await update.message.reply_text(f"✅ User `{target_id}` **granted access**.", parse_mode='Markdown')
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}")
    return ConversationHandler.END

async def admin_broadcast(update, context):
    message_text = update.message.text
    members = redis_client.smembers("auth_users")
    success = 0
    fail = 0
    await update.message.reply_text(f"📤 Sending to {len(members)} users...")
    for uid in members:
        try:
            await context.bot.send_message(
                chat_id=int(uid),
                text=f"📢 *Admin Announcement*\n\n{message_text}",
                parse_mode='Markdown'
            )
            success += 1
        except Exception:
            fail += 1
    await update.message.reply_text(
        f"✅ Broadcast complete!\n📤 Sent: {success}\n❌ Failed: {fail}"
    )
    return ConversationHandler.END

async def admin_cancel(update, context):
    if update.message:
        await update.message.reply_text("❌ Admin action cancelled.")
    return ConversationHandler.END

def main():
    async def post_init(application):
        """Auto-register bot commands so they appear in Telegram Menu button."""
        # Regular user commands
        await application.bot.set_my_commands([
            BotCommand("calculator", "Loan Amortization Calculator"),
            BotCommand("score",      "Loan Pre-approval Scoring"),
            BotCommand("valuation",  "Property Valuation Tool"),
            BotCommand("myid",       "View your Telegram ID"),
        ])
        # Admin-only commands (shown only in Admin's chat)
        if ADMIN_ID:
            await application.bot.set_my_commands([
                BotCommand("calculator", "Loan Amortization Calculator"),
                BotCommand("score",      "Loan Pre-approval Scoring"),
                BotCommand("valuation",  "Property Valuation Tool"),
                BotCommand("admin",      "👑 Admin Panel"),
                BotCommand("myid",       "View your Telegram ID"),
            ], scope=BotCommandScopeChat(chat_id=ADMIN_ID))
        print("✅ Bot commands registered with Telegram.")

    app = ApplicationBuilder().token(BOT_TOKEN).post_init(post_init).build()
    # /myid in group=-1 so it always works from any state
    app.add_handler(CommandHandler("myid", myid), group=-1)
    
    # Common handlers to allow switching between tools from any state
    common_handlers = [
        MessageHandler(filters.Regex("^🧮 Calculator$"), start_calculator),
        MessageHandler(filters.Regex("^📋 Score$"), start_score),
        MessageHandler(filters.Regex("^🏢 Valuation$"), start_valuation),
        MessageHandler(filters.Regex("^👑 Admin Panel$"), admin_panel),
    ]

    main_conv = ConversationHandler(
        entry_points=[
            CommandHandler("start", start),
            CommandHandler("calculator", start_calculator),
            CommandHandler("score", start_score),
            CommandHandler("valuation", start_valuation),
            CommandHandler("admin", admin_panel),
            *common_handlers
        ],
        states={
            # Calculator States
            METHOD: [*common_handlers, CallbackQueryHandler(select_method)],
            AMOUNT: [*common_handlers, MessageHandler(filters.TEXT & ~filters.COMMAND, get_amount)],
            RATE: [*common_handlers, MessageHandler(filters.TEXT & ~filters.COMMAND, get_rate)],
            TERM_MONTHS: [*common_handlers, MessageHandler(filters.TEXT & ~filters.COMMAND, get_term_months)],
            
            # Valuation States
            VAL_DISTRICT: [*common_handlers, CallbackQueryHandler(select_district)],
            VAL_SIZE: [*common_handlers, MessageHandler(filters.TEXT & ~filters.COMMAND, get_val_size)],
            
            # Score States
            SCORE_INCOME: [*common_handlers, MessageHandler(filters.TEXT & ~filters.COMMAND, get_score_income)],
            SCORE_DEBT: [*common_handlers, MessageHandler(filters.TEXT & ~filters.COMMAND, get_score_debt)],
            SCORE_PROP_VALUE: [*common_handlers, MessageHandler(filters.TEXT & ~filters.COMMAND, get_score_prop_value)],
            SCORE_LOAN_AMOUNT: [*common_handlers, MessageHandler(filters.TEXT & ~filters.COMMAND, get_score_loan_amount)],
            
            # Admin States
            ADMIN_MENU: [CallbackQueryHandler(admin_callback, pattern="^admin_")],
            ADMIN_REVOKE_INPUT: [CallbackQueryHandler(admin_revoke_callback, pattern="^revoke_|^admin_back$")],
            ADMIN_ADD_INPUT: [*common_handlers, MessageHandler(filters.TEXT & ~filters.COMMAND, admin_add_user)],
            ADMIN_BROADCAST_INPUT: [*common_handlers, MessageHandler(filters.TEXT & ~filters.COMMAND, admin_broadcast)],
        },
        fallbacks=[
            CommandHandler("cancel", start),
            *common_handlers
        ],
        allow_reentry=True
    )
    app.add_handler(main_conv)
    
    async def error_handler(update, context):
        import traceback
        print(f"\u274c Error: {context.error}")
        traceback.print_exc()
        if update and hasattr(update, 'message') and update.message:
            await update.message.reply_text(
                "⚠️ Something went wrong. Please try /start again."
            )
    app.add_error_handler(error_handler)
    
    app.add_handler(
        MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message)
    )
    if os.getenv("RENDER"):
        print("✅ Starting bot via Webhook on Render...")
        PORT = int(os.environ.get('PORT', 8080))
        RENDER_EXTERNAL_URL = os.environ.get('RENDER_EXTERNAL_URL', 'https://loan-bot-qyzu.onrender.com')
        app.run_webhook(
            listen="0.0.0.0",
            port=PORT,
            webhook_url=RENDER_EXTERNAL_URL
        )
    else:
        print("✅ Bot is running locally via Polling...")
        keep_alive()
        app.run_polling()

if __name__ == "__main__":
    main()